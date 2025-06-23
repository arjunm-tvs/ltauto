'''USES NETLIST PARSER AND USES RAW FILE TIME WAVEFORM'''



from pathlib import Path
from PyLTSpice import RawRead
import numpy as np
import csv
from openpyxl import Workbook

def is_component_line(line):
    skip_keywords = ['.model', '.lib', '.tran', '.TEMP', '.end', '.backanno']
    if not line or line.startswith(('*', ';')):
        return False
    return not any(line.lower().startswith(k.lower()) for k in skip_keywords)

# To identify the nodes between the devices

def parse_netlist(filepath):
    netlist_path = Path(filepath)
    components = {}

    with netlist_path.open() as f:
        for line in f:
            line = line.strip()
            if not is_component_line(line):
                continue

            tokens = line.split()
            if not tokens:
                continue

            comp_name = tokens[0]
            nodes = [t for t in tokens[1:] if t.upper().startswith("N") or t == '0']

            if len(nodes) == 1:
                nodes.append("0")

            components[comp_name] = tuple(nodes)

    return components

def analyze_power(raw_file, component_nodes):
    raw = RawRead(raw_file)
    results = {}
    
    time = raw.get_trace('time').get_wave()

    trace_names = raw.get_trace_names()
   
   # Extract node voltages
    peak_node_values = {}
    for trace in trace_names:
        if trace.startswith('V(') and trace.endswith(')'):
            try:
                voltage_wave = raw.get_trace(trace).get_wave()
                peak_voltage = np.max(np.abs(voltage_wave))
                peak_node_values[trace] = {
                    'Peak Voltage': round(peak_voltage, 6)
                }
            except Exception as e:
                print(f"⚠️ Skipping {trace} due to error: {e}")
   
    
        

    for comp, nodes in component_nodes.items():
        
        if len(nodes) != 2:
            continue

        n_plus, n_minus = nodes
        
        # print(n_plus,n_minus)

        v_plus_trace = f"V({n_plus.lower()})"
        v_minus_trace = f"V({n_minus.lower()})"
        
        # print(v_plus_trace,v_minus_trace)

        v_plus = raw.get_trace(v_plus_trace).get_wave() if v_plus_trace in trace_names else 0
        v_minus = raw.get_trace(v_minus_trace).get_wave() if v_minus_trace in trace_names else 0
        
        # print(v_plus,v_minus)
        
        voltage = v_plus - v_minus

        current_trace = f"I({comp})"
        current = raw.get_trace(current_trace).get_wave() 

        power = voltage * current

        results[comp] = {
            "peak_voltage": np.max(np.abs(voltage)),
            "peak_current": np.max(np.abs(current)),
            "peak_power": np.max(np.abs(power))
        }

    return results, peak_node_values

# def export_to_csv(results, output_csv_path):
#     with open(output_csv_path, mode='w', newline='') as csvfile:
#         writer = csv.writer(csvfile)
#         writer.writerow(['Component', 'Peak Voltage (V)', 'Peak Current (A)', 'Peak Power (W)'])

#         for comp, data in results.items():
#             writer.writerow([
#                 comp,
#                 f"{data['peak_voltage']:.6f}",
#                 f"{data['peak_current']:.6f}",
#                 f"{data['peak_power']:.6f}"
#             ])
        
#         print(f"✅  Peak analysis exported to: {output_csv_path}")
def export_to_csv(results, peak_node_values, output_csv_path):
    with open(output_csv_path, mode='w', newline='') as csvfile:
        writer = csv.writer(csvfile)

        # --- Component Power Analysis ---
        writer.writerow(['Component Power Analysis'])
        writer.writerow(['Component', 'Peak Voltage (V)', 'Peak Current (A)', 'Peak Power (W)'])

        for comp, data in results.items():
            writer.writerow([
                comp,
                f"{data['peak_voltage']:.6f}",
                f"{data['peak_current']:.6f}",
                f"{data['peak_power']:.6f}"
            ])

        writer.writerow([])  # Blank line

        # --- Node Voltage Peaks ---
        writer.writerow(['Peak Node Voltages'])
        writer.writerow(['Node', 'Peak Voltage (V)'])

        for node, data in peak_node_values.items():
            writer.writerow([node, f"{data['Peak Voltage']:.6f}"])

    print(f"✅ Combined results exported to: {output_csv_path}")



# def export_to_excel(results, output_excel_path):
#     # Create a new workbook and get the active sheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Simulation Results"

#     # Write the header
#     ws.append(['Component', 'Peak Voltage (V)', 'Peak Current (A)', 'Peak Power (W)'])

#     # Write data rows
#     for comp, data in results.items():
#         ws.append([
#             comp,
#             round(data['peak_voltage'], 6),
#             round(data['peak_current'], 6),
#             round(data['peak_power'], 6)
#         ])

#     # Save the workbook
#     wb.save(output_excel_path)
#     print(f"✅ Excel file saved at: {output_excel_path}")

from openpyxl import Workbook

def export_to_excel(results, peak_node_values, output_excel_path):
    wb = Workbook()

    # --- Sheet 1: Component Power Analysis ---
    ws1 = wb.active
    ws1.title = "Component Power"

    # Header
    ws1.append(['Component', 'Peak Voltage (V)', 'Peak Current (A)', 'Peak Power (W)'])

    for comp, data in results.items():
        ws1.append([
            comp,
            round(data['peak_voltage'], 6),
            round(data['peak_current'], 6),
            round(data['peak_power'], 6)
        ])

    # --- Sheet 2: Node Peak Voltages ---
    ws2 = wb.create_sheet(title="Node Peaks")
    ws2.append(['Node', 'Peak Voltage (V)'])

    for node, data in peak_node_values.items():
        ws2.append([node, round(data['Peak Voltage'], 6)])

    # Save file
    wb.save(output_excel_path)
    print(f"✅ Combined results in Excel file saved at: {output_excel_path}")



# --- File Paths ---

# raw_file = r"D:\OneDrive - TVS Motor Company Ltd\Desktop\simdata1\FPL_center_1.raw"
# csv_output = r"D:\OneDrive - TVS Motor Company Ltd\Desktop\s imdata1\exported\peak_analysis_output.csv"



netlist_path_general = Path(r"D:\OneDrive - TVS Motor Company Ltd\Desktop\simdata1")

for temp in [-10, 25, 85]:
    
    netlist_temp_general = netlist_path_general / f'TEMP_{temp}'
    
    for voltage_level in [9, 12, 16]:

        # --- File Paths ---
        netlist_path_exact = netlist_temp_general / f'Voltage_level_{voltage_level}' / "FPl_CENTER_1_1.net"
        raw_file = netlist_temp_general / f'Voltage_level_{voltage_level}' / "FPl_CENTER_1_1.raw"
        csv_output = netlist_temp_general / f'Peak_values_voltage_level_{voltage_level}_csv.csv'
        excel_output = netlist_temp_general / f'Peak_values_voltage_level_{voltage_level}_excel.xlsx'

        if netlist_path_exact.exists() and raw_file.exists():
            print(f"\nProcessing TEMP={temp}, Voltage={voltage_level}...")
            
            component_nodes = parse_netlist(netlist_path_exact)
            results,peak_node_values = analyze_power(raw_file, component_nodes)

            export_to_csv(results,peak_node_values, csv_output)
            export_to_excel(results,peak_node_values, excel_output)
        else:
            print(f"⚠️  Missing files for TEMP={temp}, Voltage={voltage_level}")
            print(f"  Netlist: {netlist_path_exact.exists()}, Raw: {raw_file.exists()}")

        